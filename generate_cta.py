#!/usr/bin/env python3
"""
CTA未挿入の記事に対して、Gemini APIで最適な資料を選定しCTAコラムHTMLを生成する。

ワークフロー:
  1. Excelから「記事別CTRCVR管理」タブのフォーム1〜5が空欄の記事を抽出
  2. 各記事URLにアクセスして本文を取得
  3. 資料一覧と照合して最適な1〜3点を選定
  4. CTAコラム文言・HTMLを生成
  5. 結果をExcelとJSONに保存

使い方:
  python generate_cta.py                       # 全未挿入記事を処理
  python generate_cta.py --limit 5             # 最大5件
  python generate_cta.py --slug 0116           # 特定記事のみ
"""

import os
import sys
import json
import re
import csv
import time
import argparse
from datetime import datetime
from pathlib import Path

import requests
from bs4 import BeautifulSoup
import openpyxl

try:
    from google import genai
    from google.genai import types
except ImportError:
    print("google-genai がインストールされていません。")
    sys.exit(1)

# ── 設定 ─────────────────────────────────────────────────
API_KEY = os.environ.get("GEMINI_API_KEY")
if not API_KEY:
    print("環境変数 GEMINI_API_KEY が設定されていません。")
    sys.exit(1)

MODEL = os.environ.get("GEMINI_MODEL", "gemini-2.5-flash")
DIR = Path(__file__).resolve().parent
OUTPUT_DIR = DIR / "cta_output"
OUTPUT_DIR.mkdir(exist_ok=True)

EXCEL_PATH = Path.home() / "Desktop" / "1on1総研_記事別挿入リンク.xlsx"
PV_CSV_PATH = Path.home() / "Desktop" / "KAKEAI様_BI_1on1総研_表.csv"
MEDIA_BASE = "https://kakeai.co.jp"

# ── 資料一覧 ─────────────────────────────────────────────
RESOURCES = [
    {
        "id": 1,
        "title": "部門別施策別1on1最新事例",
        "url": "https://clw4m.share-na2.hsforms.com/2BmKkXDojRjyWA3Q-gf3EKg",
        "summary": "部門特性・施策別の1on1定着・質向上事例",
    },
    {
        "id": 2,
        "title": "現場の対話が組織を革新する（KDDI SF）",
        "url": "https://clw4m.share-na2.hsforms.com/2Zb9mGwq2TKuHqGc6twewOg",
        "summary": "KDDI Sonic-Falcon事例、二層式1on1・95%実施率・離職防止と育成の両立",
    },
    {
        "id": 3,
        "title": "0から始める！1on1導入ガイドライン",
        "url": "https://clw4m.share-na2.hsforms.com/2IUfK3sI8T-WdkU7kTH95Pw",
        "summary": "目的明確化→ルール策定→モニタリングまで、準備・運用・定着の3ステップ",
    },
    {
        "id": 4,
        "title": "マネジャーのための1on1完全ガイド",
        "url": "https://clw4m.share-na2.hsforms.com/2yKshXPStTWm7iktGvbSRlg",
        "summary": "準備・実施・振り返り・FAQ・質問テンプレート（6カテゴリー）",
    },
    {
        "id": 5,
        "title": "エンゲージメント向上につながる1on1とは",
        "url": "https://clw4m.share-na2.hsforms.com/21ewyUerBS2CCxh4OVtO06Q",
        "summary": "導入企業の成功事例・効果",
    },
    {
        "id": 6,
        "title": "心理的安全性を高める組織づくりと1on1の実践",
        "url": "https://clw4m.share-na2.hsforms.com/2ygReXW3vTtO9w5waBIijiQ",
        "summary": "静かな壁の正体・段階別（低・中・高）設計・運用",
    },
    {
        "id": 7,
        "title": "メンバーのための1on1完全ガイド",
        "url": "https://clw4m.share-na2.hsforms.com/2eGL37SJ9QwqhBD7x-So9fw",
        "summary": "メンバーが1on1を自分の時間として活用する方法・テーマの決め方・振り返り",
    },
    {
        "id": 8,
        "title": "メンター・横・斜めとの1on1",
        "url": "https://clw4m.share-na2.hsforms.com/2iNXAO94uRhm4dZmYafUARg",
        "summary": "多対話構造・斜めの1on1の設計・注目される背景",
    },
    {
        "id": 9,
        "title": "マネジャー向け1on1質問集",
        "url": "https://clw4m.share-na2.hsforms.com/28M4ZFIxaTmC8y6F0sEhVlQ",
        "summary": "タイプ別（新人・若手・年上・リーダー候補）質問集",
    },
    {
        "id": 10,
        "title": "1on1関係性サーベイ・16段階",
        "url": "https://clw4m.share-na2.hsforms.com/2vs3-L-LjRxiskpoHMoXmPQ",
        "summary": "270万回データから16段階の関係性モデル・部下が回答するサーベイ形式",
    },
    {
        "id": 11,
        "title": "1on1お役立ち資料3点セット",
        "url": "https://clw4m.share-na2.hsforms.com/2wDGox01zS8mKOr0hdIOduw",
        "summary": "複数資料のセットダウンロード",
    },
]


# ── PVデータ読み込み ─────────────────────────────────────
def load_pv_data(csv_path):
    """CSVから記事URLごとの表示回数を辞書で返す"""
    pv_map = {}
    if not csv_path.exists():
        print(f"  警告: PV CSVが見つかりません: {csv_path}")
        return pv_map

    with open(csv_path, encoding="utf-8") as f:
        reader = csv.reader(f)
        headers = next(reader)
        # 「表示回数」列のインデックスを探す
        pv_col = None
        url_col = None
        for i, h in enumerate(headers):
            if "表示回数" in h:
                pv_col = i
            if "URL" in h or "パス" in h:
                url_col = i
        if pv_col is None or url_col is None:
            print(f"  警告: CSV列が見つかりません (表示回数={pv_col}, URL={url_col})")
            return pv_map

        for row in reader:
            if len(row) <= max(pv_col, url_col):
                continue
            url = row[url_col].strip()
            try:
                pv = int(row[pv_col].replace(",", ""))
            except (ValueError, IndexError):
                pv = 0
            pv_map[url] = pv

    return pv_map


# ── Step 1: CTA未挿入記事を抽出（PV順ソート対応） ────────
def load_empty_articles(excel_path, pv_map=None, slug_filter=None, limit=None):
    """Excelからフォーム1〜5が空欄の記事を抽出し、PV降順でソート"""
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    ws = wb["記事別CTRCVR管理 "]

    articles = []
    for r in range(2, ws.max_row + 1):
        title = ws.cell(r, 3).value  # C: 記事タイトル
        if not title:
            continue

        slug = ws.cell(r, 2).value   # B: slug
        url = ws.cell(r, 4).value    # D: 記事URL

        # フォーム1〜5 (J〜N = col 10〜14)
        forms = [ws.cell(r, c).value for c in range(10, 15)]
        has_cta = any(v is not None and v != "" for v in forms)

        if has_cta:
            continue

        if slug_filter and str(slug) != slug_filter:
            continue

        # ダミー記事を除外
        if "ダミー" in str(title):
            continue

        # PV数を付与
        url_str = str(url) if url else ""
        pv = pv_map.get(url_str, 0) if pv_map else 0

        articles.append({
            "row": r,
            "slug": str(slug),
            "title": str(title),
            "url": url_str,
            "pv": pv,
        })

    wb.close()

    # PV降順でソート
    articles.sort(key=lambda a: a["pv"], reverse=True)

    # limit適用
    if limit:
        articles = articles[:limit]

    return articles


# ── Step 2: 記事本文を取得 ────────────────────────────────
# CTA存在を示すマーカー（ページ上にこれらがあれば挿入済みと判定）
CTA_MARKERS = ["📂", "📘", "\U0001F4C2", "\U0001F4D8"]


def fetch_article_body(article_url):
    """記事ページから本文テキストを取得。戻り値は (text, has_cta_on_page) のタプル。"""
    full_url = f"{MEDIA_BASE}{article_url}" if article_url.startswith("/") else article_url
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) 1on1-cta-bot"
    }

    try:
        resp = requests.get(full_url, headers=headers, timeout=30)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")

        # 記事本文エリアを探す
        body = soup.find("article") or soup.find(class_=re.compile(r"article|entry|post|content"))
        if body:
            # scriptやstyleを除去
            for tag in body.find_all(["script", "style", "nav", "footer"]):
                tag.decompose()
            text = body.get_text(separator="\n", strip=True)
        else:
            text = soup.get_text(separator="\n", strip=True)

        # CTA挿入済み判定: ページ全体のHTMLテキストでマーカーを検索
        page_text = soup.get_text(separator=" ", strip=True)
        has_cta = any(marker in page_text for marker in CTA_MARKERS)

        # 先頭6000文字に制限（トークン節約）
        return text[:6000], has_cta
    except Exception as e:
        return f"(取得失敗: {e})", False


# ── Step 3-4: Gemini APIで資料選定 + CTA文言生成 ─────────
def generate_cta_for_article(client, article, body_text):
    """1記事に対してCTA資料選定 + コラム文言生成"""

    resources_text = "\n".join(
        f"  {r['id']}. {r['title']} — {r['summary']}\n     URL: {r['url']}"
        for r in RESOURCES
    )

    prompt = f"""あなたはBtoBコンテンツマーケティングのCRO（コンバージョン率最適化）専門家です。
1on1総研のオウンドメディア記事に、読者の行動を自然に促すCTAコラムを設計してください。

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
■ あなたの専門知識（CTA設計の原則）
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

【配置戦略】BtoBオウンドメディアの最適CTA配置（Cone社調査に基づく）
- 位置A「リード文直下 / 目次上」: 読者が"価値がありそう"と判断した直後。情報収集層を早期に捕捉
- 位置B「2番目のH2見出し直前」: 読了率が低下し始める離脱直前。ここを逃すと読者の半数以上を失う
- 位置C「本文末尾」: 最後まで読んだ高関心層。最もCVRが高い位置
→ 1〜3点のCTAをこの3箇所に分散配置する。同じ箇所に2つ以上置かない。

【文言設計】文言改善は配置改善の2倍の効果（商談化率30%以上達成: 文言57.2% vs 配置31.2%）
- PAS構造を使う: Problem（読者の課題を代弁）→ Agitation（放置リスクを示唆）→ Solution（資料で解決）
- 得られる価値を具体的に書く: ×「〜について解説」→ ○「〜を実現するための3ステップがわかる」
- 行動動詞で終わる: 「手に入れる」「確認する」「始める」
- 「無料」を明示して心理的ハードルを下げる（採用率62.2%の最重要トリガー）
- 記事の文脈から自然に橋渡しする（文脈連動CTAは汎用CTAの3〜10倍のCVR）

【読者フェーズの見極め】
- 情報収集層（42.2%）: 知識欲を満たす資料（ガイド・レポート・チェックリスト）
- 課題解決層（47.6%）: 自社の課題に合致する具体的手法（事例・テンプレート）
- 即アクション層（34.4%）: サービス資料・相談

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
■ 対象記事
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
- タイトル: {article['title']}
- URL: {MEDIA_BASE}{article['url']}

## 記事本文（抜粋）
{body_text[:5000]}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
■ 資料一覧（全11件を必ずフラットに照合すること）
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{resources_text}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
■ 設計プロセス（必ずこの順序で思考すること）
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

### STEP 1: 記事の読者像とニーズを分析
- この記事を読む人はどんな課題を抱えているか？
- 読了後にどんな「次の行動」を取りたくなるか？
- 情報収集層 / 課題解決層 / 即アクション層、どの比率が高いか？

### STEP 2: 資料を全件照合し、読者ニーズとの適合度を判定
- 全11件について、記事読者の課題・関心との適合度を判定
- 「マネジャーのための1on1完全ガイド」等の汎用資料を安易に選ばない
- 記事テーマに"直接的に"答える資料を最優先する

### STEP 3: 1〜3点を選定し、配置位置を決定
- 関連性が薄ければ無理に3点選ばない（1点でもOK）
- 各CTAを位置A/B/Cのいずれかに配置し、同じ位置に2つ置かない
- 各位置でターゲットする読者フェーズを明記する

### STEP 4: 記事文脈からの橋渡し文を設計
- 挿入箇所の直前の本文を引用し、その文脈を受けて自然にCTAに入る
- 唐突な資料紹介にならないよう、記事の論点→読者の課題→資料の価値、の流れを作る

### STEP 5: PAS構造でCTAコラム文言を生成
各CTAについて以下のHTMLを生成：

```
<div style="background:#f5f5f5; border-left:4px solid #4a4a8a; padding:16px 20px; margin:24px 0;">
  <p style="font-weight:bold; margin:0 0 8px;">📂 コラムタイトル（読者の課題を反映した具体的なタイトル）</p>
  <p style="margin:0 0 16px;">PAS構造の説明文:
    Problem — 記事で触れた課題を読者視点で代弁（「〜に悩んでいませんか？」ではなく、断定的に課題を描写）
    → Agitation — その課題を放置するとどうなるかを1文で示唆
    → Solution — この資料で何が得られるかを具体的に（「3つのステップ」「実践テンプレート」等）
  </p>
  <p style="margin:0;">📘 <a href="フォームURL" target="_blank">資料タイトル（無料）</a></p>
</div>
```

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
■ 品質チェックリスト（生成後に必ず自己検証）
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
□ 記事の文脈から自然に橋渡しされているか？（唐突に「下記の資料では〜」と始まっていないか）
□ 読者が「これは自分のための情報だ」と感じる文言になっているか？
□ 資料で得られる価値が具体的に書かれているか？（「〜を解説」で終わっていないか）
□ 「無料」が明示されているか？
□ 各CTAの挿入位置が分散しているか？（同じ箇所に集中していないか）
□ 汎用的すぎる資料選定になっていないか？（記事テーマに直接答える資料か）

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
■ 禁止事項
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
- 資料の中身を推測で書かない（概要に書かれた情報のみ使用）
- 「こちら」「ぜひ」「こちらもご覧ください」等の広告的表現は使わない
- 本文の途中（段落の途中）に挿入しない
- 「〜について解説しています」だけで終わる説明文は禁止（得られる具体的価値を書く）
- 同じ挿入箇所に複数のCTAを配置しない

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
■ 出力形式（JSON）
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

```json
{{
  "article_theme": "記事テーマの一言定義",
  "target_reader": "想定読者像（例: 初めて1on1を導入する人事担当者）",
  "reader_phase": "主な読者フェーズ（情報収集/課題解決/即アクション）",
  "ctas": [
    {{
      "resource_id": 数字,
      "resource_title": "資料タイトル",
      "resource_url": "フォームURL",
      "placement": "A/B/Cのいずれか",
      "placement_rationale": "この位置に置く理由（読者の心理状態を踏まえて）",
      "insertion_point": "挿入箇所の直前の本文一文を引用",
      "target_phase": "このCTAがターゲットする読者フェーズ",
      "column_title": "コラムのタイトル（📂の後に表示）",
      "column_text": "PAS構造の説明文",
      "html": "完全なHTMLコード"
    }}
  ]
}}
```
"""

    max_retries = 5
    for attempt in range(max_retries):
        try:
            response = client.models.generate_content(
                model=MODEL,
                contents=prompt,
                config=types.GenerateContentConfig(
                    temperature=0.4,
                    max_output_tokens=4000,
                    response_mime_type="application/json",
                ),
            )
            raw = response.text.strip()
            return _parse_json_response(raw)
        except Exception as e:
            err_str = str(e)
            if "429" in err_str or "RESOURCE_EXHAUSTED" in err_str:
                # retryDelay をパース（例: "retryDelay": "14s"）
                delay_match = re.search(r'"retryDelay":\s*"(\d+)', err_str)
                wait_sec = int(delay_match.group(1)) + 5 if delay_match else 30
                wait_sec = max(wait_sec, 15)  # 最低15秒
                if attempt < max_retries - 1:
                    print(f"  ⏳ レート制限: {wait_sec}秒待機してリトライ ({attempt+1}/{max_retries})...")
                    time.sleep(wait_sec)
                    continue
            return {"article_theme": f"エラー: {e}", "ctas": []}


def _parse_json_response(raw):
    """GeminiのJSON出力をパース。HTML含有で壊れやすいので複数戦略で試行"""
    # 1. そのままパース
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        pass

    # 2. 全角引用符を半角に置換
    cleaned = raw.replace("\u201c", '"').replace("\u201d", '"')
    cleaned = cleaned.replace("\u2018", "'").replace("\u2019", "'")
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        pass

    # 3. htmlフィールドを丸ごと空にしてパースし、後でregexで復元
    stripped = re.sub(r'"html"\s*:\s*".*?"(\s*[,\}])', r'"html": ""\1', cleaned, flags=re.DOTALL)
    try:
        result = json.loads(stripped)
        html_blocks = re.findall(r'<div style=.*?</div>', raw, re.DOTALL)
        ctas = result.get("ctas", [])
        for i, cta in enumerate(ctas):
            if i < len(html_blocks):
                cta["html"] = html_blocks[i]
        return result
    except json.JSONDecodeError:
        pass

    # 4. 新戦略: エスケープされたHTMLを含む場合に対応（\\\"→"変換後にHTMLを抽出）
    unescaped = cleaned.replace('\\"', '"').replace('\\n', '\n')
    html_blocks = re.findall(r'<div style=.*?</div>', unescaped, re.DOTALL)

    # JSONフィールドを個別にregexで抽出
    def _extract(field, text):
        m = re.search(rf'"{field}"\s*:\s*"((?:[^"\\]|\\.)*)"', text)
        return m.group(1).replace('\\"', '"').replace('\\n', '\n') if m else ""

    theme = _extract("article_theme", cleaned)
    target_reader = _extract("target_reader", cleaned)
    reader_phase = _extract("reader_phase", cleaned)

    # 各CTAオブジェクトをregexで抽出
    cta_blocks = re.findall(r'\{[^{}]*"resource_id"[^{}]*\}', cleaned, re.DOTALL)
    ctas = []
    for idx, block in enumerate(cta_blocks):
        cta = {
            "resource_id": int(m.group(1)) if (m := re.search(r'"resource_id"\s*:\s*(\d+)', block)) else 0,
            "resource_title": _extract("resource_title", block),
            "resource_url": _extract("resource_url", block),
            "placement": _extract("placement", block),
            "placement_rationale": _extract("placement_rationale", block),
            "insertion_point": _extract("insertion_point", block),
            "target_phase": _extract("target_phase", block),
            "column_title": _extract("column_title", block),
            "column_text": _extract("column_text", block),
            "html": html_blocks[idx] if idx < len(html_blocks) else "",
        }
        ctas.append(cta)

    if ctas or theme:
        return {
            "article_theme": theme,
            "target_reader": target_reader,
            "reader_phase": reader_phase,
            "ctas": ctas,
            "_parse_note": "regex抽出によるフォールバックパース",
        }

    # 5. 最終フォールバック
    return {
        "article_theme": theme or "パース失敗",
        "ctas": [{"html": h, "resource_title": "要確認", "resource_url": "", "insertion_point": "",
                  "column_title": "", "column_text": "", "placement": "", "placement_rationale": "",
                  "target_phase": ""} for h in html_blocks],
        "_parse_note": "最終フォールバック: HTMLブロックのみ抽出",
    }


# ── Step 5: 結果を保存 ──────────────────────────────────
def save_results(results, output_dir):
    """JSON + テキストで保存"""
    date_str = datetime.now().strftime("%Y-%m-%d_%H%M")

    # JSON保存
    json_path = output_dir / f"{date_str}_cta_results.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"\n  JSON保存: {json_path}")

    # テキスト保存（レビュー用）
    txt_path = output_dir / f"{date_str}_cta_results.txt"
    lines = [f"# CTA生成結果 ({date_str})", f"# 処理件数: {len(results)}件", ""]

    for r in results:
        lines.append("=" * 70)
        lines.append(f"記事: [{r['slug']}] {r['title']}")
        lines.append(f"URL: {MEDIA_BASE}{r['url']}")
        lines.append(f"PV: {r.get('pv', 0)}")
        lines.append(f"テーマ: {r.get('article_theme', '?')}")
        lines.append(f"想定読者: {r.get('target_reader', '?')}")
        lines.append(f"読者フェーズ: {r.get('reader_phase', '?')}")
        lines.append(f"CTA数: {len(r.get('ctas', []))}件")
        lines.append("")

        for i, cta in enumerate(r.get("ctas", []), 1):
            lines.append(f"--- CTA {i} [位置{cta.get('placement', '?')}] ---")
            lines.append(f"資料: {cta.get('resource_title', '?')}")
            lines.append(f"配置理由: {cta.get('placement_rationale', '?')}")
            lines.append(f"対象フェーズ: {cta.get('target_phase', '?')}")
            lines.append(f"挿入箇所: {cta.get('insertion_point', '?')[:100]}")
            lines.append(f"コラムタイトル: {cta.get('column_title', '?')}")
            lines.append(f"説明文: {cta.get('column_text', '?')}")
            lines.append(f"HTML:")
            lines.append(cta.get("html", ""))
            lines.append("")

        lines.append("")

    txt_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"  テキスト保存: {txt_path}")

    return json_path, txt_path


# ── メイン ──────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="CTA一括生成スクリプト")
    parser.add_argument("--limit", type=int, default=None, help="処理する最大記事数")
    parser.add_argument("--slug", type=str, default=None, help="特定のslugのみ処理")
    parser.add_argument("--excel", type=str, default=None, help="Excelファイルパス")
    parser.add_argument("--csv", type=str, default=None, help="PV CSVファイルパス")
    args = parser.parse_args()

    excel_path = Path(args.excel) if args.excel else EXCEL_PATH
    csv_path = Path(args.csv) if args.csv else PV_CSV_PATH

    if not excel_path.exists():
        print(f"エラー: Excelファイルが見つかりません: {excel_path}")
        sys.exit(1)

    print(f"===== CTA一括生成 =====")
    print(f"  Excel: {excel_path}")
    print(f"  PV CSV: {csv_path}")

    # PVデータ読み込み
    print("\n[1/5] PVデータを読み込み...")
    pv_map = load_pv_data(csv_path)
    print(f"  → {len(pv_map)} 件のPVデータ")

    # CTA未挿入記事を抽出（PV順ソート）
    print("\n[2/5] CTA未挿入記事を抽出（PV降順）...")
    articles = load_empty_articles(excel_path, pv_map=pv_map, slug_filter=args.slug, limit=args.limit)
    print(f"  → {len(articles)} 件の未挿入記事")
    for a in articles[:10]:
        print(f"    PV {a['pv']:>5} | [{a['slug']}] {a['title'][:45]}")

    if not articles:
        print("  処理する記事がありません。")
        return

    client = genai.Client(api_key=API_KEY)
    results = []

    for i, article in enumerate(articles, 1):
        print(f"\n[3-5] ({i}/{len(articles)}) [{article['slug']}] PV={article.get('pv',0)} {article['title'][:40]}...")

        # Step 2: 記事本文を取得 & CTA挿入済みチェック
        print("  本文を取得中...")
        body, has_cta_on_page = fetch_article_body(article["url"])
        if isinstance(body, str) and body.startswith("(取得失敗"):
            print(f"  {body}")
            results.append({**article, "article_theme": "本文取得失敗", "ctas": []})
            continue

        if has_cta_on_page:
            print(f"  ⏭ スキップ: ページ上にCTA（📂/📘）が既に存在")
            results.append({**article, "article_theme": "CTA挿入済み（ページ上検出）", "ctas": [], "skipped": True})
            continue

        print(f"  → {len(body)} 文字取得")

        # Step 3-4: CTA生成
        print("  CTA生成中...")
        result = generate_cta_for_article(client, article, body)
        result.update(article)
        results.append(result)

        cta_count = len(result.get("ctas", []))
        theme = result.get("article_theme", "?")
        print(f"  → テーマ: {theme}")
        print(f"  → CTA {cta_count}件生成")
        for cta in result.get("ctas", []):
            print(f"    - {cta.get('resource_title', '?')}")

        # レート制限対策
        if i < len(articles):
            time.sleep(1)

    # Step 5: 結果保存
    print("\n[保存]")
    save_results(results, OUTPUT_DIR)

    # サマリー
    total_ctas = sum(len(r.get("ctas", [])) for r in results)
    skipped_page = sum(1 for r in results if r.get("skipped"))
    skipped_fetch = sum(1 for r in results if r.get("article_theme") == "本文取得失敗")
    generated = len(results) - skipped_page - skipped_fetch
    print(f"\n===== 完了 =====")
    print(f"  対象記事数: {len(results)}")
    print(f"  スキップ（ページ上CTA検出）: {skipped_page}")
    print(f"  スキップ（本文取得失敗）: {skipped_fetch}")
    print(f"  CTA生成済み: {generated}")
    print(f"  生成CTA数: {total_ctas}")
    print(f"  出力先: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
